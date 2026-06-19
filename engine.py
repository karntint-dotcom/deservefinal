"""
engine.py — Deserve Dashboard (Pure Python edition)
ทุก sheet คำนวณใน Python แล้ว hardcode ลง Excel
Sheet 6, 7, 8, 9, 11 มีคอลัมน์จำนวนชิ้น (unit) ที่ขายออก
"""
import io, warnings
import pandas as pd
import numpy as np
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter as gc
warnings.filterwarnings('ignore')

# ─── STYLES ──────────────────────────────────────────────────────────────────
T  = Side(style='thin',   color='D0D0D0')
M  = Side(style='medium', color='595959')
BT = Border(left=T, right=T, top=T, bottom=T)

def F(h): return PatternFill("solid", fgColor=h)

def W(ws, r, c, val, bg=None, fg="000000", b=False, sz=9, fmt=None,
      h="center", wrap=False, border=BT, italic=False, merge_to=None):
    cc = ws.cell(row=r, column=c)
    cc.value = val
    cc.font = Font(name="Arial", bold=b, size=sz, color=fg, italic=italic)
    cc.alignment = Alignment(horizontal=h, vertical="center", wrap_text=wrap)
    cc.border = border
    if bg:  cc.fill = F(bg)
    if fmt: cc.number_format = fmt
    if merge_to:
        ws.merge_cells(start_row=r, start_column=c, end_row=r, end_column=merge_to)
    return cc

MTH = {1:'ม.ค.',2:'ก.พ.',3:'มี.ค.',4:'เม.ย.',5:'พ.ค.',
       6:'มิ.ย.',7:'ก.ค.',8:'ส.ค.',9:'ก.ย.',10:'ต.ค.',11:'พ.ย.',12:'ธ.ค.'}
CAT_CLR = {'Dog Food':'D6E4F7','Cat Food':'E8D5F5','Supplement':'D9F0E0',
           'Healthy Snack':'FFF2CC','Deserve Life':'FFE6D9',
           'RAW MATERIAL':'F2F2F2','Main SKU':'F2F2F2'}
STORE_CAT = {
    '🌟 STAR':       {'bg':'FFF0A0','fg':'7B5E00'},
    '📈 GROWTH':     {'bg':'C6EFCE','fg':'375623'},
    '🔄 RECOVERING': {'bg':'DEEAF1','fg':'1F4E79'},
    '📉 DECLINING':  {'bg':'FCE4D6','fg':'843C0C'},
    '⚠️ WARNING':    {'bg':'FFD7D7','fg':'C00000'},
    '😴 STABLE':     {'bg':'F2F2F2','fg':'595959'},
    '🔵 SPORADIC':   {'bg':'EBF3FB','fg':'1F4E79'},
    '⚡ INACTIVE':   {'bg':'F5F5F5','fg':'AAAAAA'},
}
PAIR_STYLE = {
    'GROWING':  ('C6EFCE','375623','📈 กำลังโต'),
    'STABLE':   ('F2F2F2','595959','➡️ คงที่'),
    'DECLINING':('FCE4D6','843C0C','📉 ลดลง'),
    'DROPPED':  ('FFD7D7','C00000','⛔ หยุดซื้อ'),
    'STOPPED':  ('FFE0B2','7B3F00','🔴 หายไป'),
    'SPORADIC': ('EBF3FB','1F4E79','💤 ไม่สม่ำเสมอ'),
    'ONE_TIME': ('FAFAFA','AAAAAA','1️⃣ ครั้งเดียว'),
}

# ─── DATA LOADING & PREP ─────────────────────────────────────────────────────
def parse_monthly_targets_from_cond3(c3):
    """อ่าน monthly_targets จาก sheet เงื่อนไข3"""
    MONTH_MAP = {
        'jan':1,'feb':2,'mar':3,'apr':4,'may':5,'jun':6,
        'jul':7,'aug':8,'sep':9,'oct':10,'nov':11,'dec':12
    }
    targets = {}
    for _, row in c3.iterrows():
        key = str(row[0]).strip().lower() if pd.notna(row[0]) else ''
        val = row[1] if len(row) > 1 else None
        if key in MONTH_MAP and pd.notna(val):
            try:
                v = float(str(val).replace(',',''))
                if v > 0: targets[MONTH_MAP[key]] = v
            except: pass
    return targets

def load_data(file_bytes):
    buf = io.BytesIO(file_bytes)

    # Auto-detect data sheet: รองรับ "รายการขาย" และ "ข้อมูล*"
    xl = pd.ExcelFile(buf)
    sheet_names = xl.sheet_names
    data_sheet = sheet_names[0]
    for s in sheet_names:
        if s.startswith('รายการขาย') or s.startswith('ข้อมูล'):
            data_sheet = s; break

    buf.seek(0)
    df = pd.read_excel(buf, sheet_name=data_sheet, header=1)

    # เงื่อนไข1, เงื่อนไข2
    buf.seek(0)
    c1_sheet = next((s for s in sheet_names if 'เงื่อนไข1' in s), sheet_names[1] if len(sheet_names)>1 else sheet_names[0])
    c2_sheet = next((s for s in sheet_names if 'เงื่อนไข2' in s), sheet_names[2] if len(sheet_names)>2 else sheet_names[0])
    c1 = pd.read_excel(buf, sheet_name=c1_sheet, header=None)
    buf.seek(0); c2 = pd.read_excel(buf, sheet_name=c2_sheet, header=None)

    # เงื่อนไข3 (เป้ารายเดือน) — optional
    buf.seek(0)
    c3_sheet = next((s for s in sheet_names if 'เงื่อนไข3' in s), None)
    monthly_targets_from_file = {}
    if c3_sheet:
        c3 = pd.read_excel(buf, sheet_name=c3_sheet, header=None)
        monthly_targets_from_file = parse_monthly_targets_from_cond3(c3)

    # ── กลับไปเริ่มอ่านข้อมูลหลัก ──────────────────────────────────────────────
    buf.seek(0); df = pd.read_excel(buf, sheet_name=data_sheet, header=1)
    df = df[df['สถานะรายการ'] == 'สำเร็จ'].copy()
    for col in ['ชื่อลูกค้า','รหัสลูกค้า','รหัสสินค้า','ชื่อสินค้า','หมวดหมู่']:
        if col in df.columns: df[col] = df[col].astype(str).str.strip()
    df['ราคารวม'] = pd.to_numeric(df['ราคารวม'], errors='coerce').fillna(0)
    df['จำนวน']   = pd.to_numeric(df['จำนวน'],   errors='coerce').fillna(0)
    df['date']    = pd.to_datetime(df['วันที่ทำรายการ'], dayfirst=True, errors='coerce')
    df['month']   = df['date'].dt.month

    # Cond1
    n2s = {}; cur = None
    for _, r_ in c1.iterrows():
        s = str(r_[0]).strip() if pd.notna(r_[0]) else ''
        c = str(r_[2]).strip() if pd.notna(r_[2]) else ''
        if s and s not in ('เงื่อนไข1','ชื่อร้าน','nan'): cur = s
        if c and c not in ('ชื่อลูกค้า','nan') and cur: n2s[c] = cur
    df['ชื่อร้าน'] = df['ชื่อลูกค้า'].map(n2s).fillna(df['ชื่อลูกค้า'])

    # Cond2
    cm = {}
    for _, r_ in c2.iterrows():
        a = str(r_[2]).strip() if pd.notna(r_[2]) else ''
        b = str(r_[3]).strip() if pd.notna(r_[3]) else ''
        if b and b != 'nan' and a and a != 'nan': cm[b] = a
    df['รหัสลูกค้า_norm'] = df['รหัสลูกค้า'].map(lambda x: cm.get(x, x))

    # Product grouping
    # กลุ่มที่ต้องแยกตามรหัสเต็ม (คนละสินค้าจริงๆ แม้ 6 ตัวแรกเหมือนกัน)
    SPLIT_GROUPS = {'100012', 'DIC130'}
    # DIC1304 มี typo 2 ชื่อ — normalize ก่อน
    NAME_FIX = {
        'Deserve ice cream (puppies and rainbow) ผักโขม เบคคอน':
        'Deserve ice cream (puppies and rainbow) ผักโขม เบคอน',
    }
    df['ชื่อสินค้า'] = df['ชื่อสินค้า'].map(lambda x: NAME_FIX.get(x, x))

    def assign_grp(code):
        prefix = str(code)[:6]
        if prefix in SPLIT_GROUPS:
            return str(code)  # ใช้รหัสเต็มเป็น group key
        return prefix

    df['prod_grp'] = df['รหัสสินค้า'].apply(assign_grp)

    prod_names = {}; prod_cat = {}
    for grp, gdf in df.groupby('prod_grp'):
        nm = gdf['ชื่อสินค้า'].value_counts()
        for n in nm.index:
            if not any(ch.isdigit() for ch in str(n)[-8:]): prod_names[grp]=n; break
        if grp not in prod_names: prod_names[grp] = nm.index[0]
        prod_cat[grp] = gdf['หมวดหมู่'].mode()[0]

    return df, prod_names, prod_cat, monthly_targets_from_file

def pivot(df, grp, months, val='ราคารวม'):
    pm = df.groupby([grp,'month'])[val].sum().unstack(fill_value=0)
    for m in months:
        if m not in pm.columns: pm[m] = 0
    return pm[months]

def _linreg(x, y):
    xm,ym = x.mean(),y.mean()
    ss_xy = ((x-xm)*(y-ym)).sum(); ss_xx = ((x-xm)**2).sum()
    sl = ss_xy/ss_xx if ss_xx != 0 else 0.0; ic = ym-sl*xm
    ss_tot = ((y-ym)**2).sum()
    r2 = max(0.0, 1-((y-(sl*x+ic))**2).sum()/ss_tot) if ss_tot != 0 else 0.0
    return sl, ic, r2

def make_forecast(vals, months_act, months_fc):
    x = np.array(months_act, dtype=float)
    sl, ic, r2 = _linreg(x, vals)
    fc = {m: max(0.0, sl*m+ic) for m in months_fc}
    wma = float(max(0, np.average(vals[-3:], weights=[1,2,3])))
    return sl, r2, fc, wma

# ════════════════════════════════════════════════════════════════════════════
# SHEET 1: ภาพรวมยอดขาย
# ════════════════════════════════════════════════════════════════════════════
def build_sheet1(wb, df, grand_total, months, monthly_targets=None):
    ws = wb.create_sheet("1_ภาพรวมยอดขาย"); ws.sheet_properties.tabColor = "1F4E79"
    mt = monthly_targets or {}
    total_target = sum(mt.values()) if mt else 0
    overall_ach  = grand_total/total_target if total_target else None

    W(ws,1,1,"Deserve Dashboard – ภาพรวมยอดขาย",bg="1F4E79",fg="FFFFFF",b=True,sz=14,h="left",merge_to=8)
    ws.row_dimensions[1].height = 26

    # KPI row — เพิ่ม Target + % Achievement
    kpis = [("ยอดขายรวม (บาท)", grand_total, '#,##0.00'),
            ("จำนวนรายการ", len(df), '#,##0'),
            ("จำนวนชื่อร้าน", df['ชื่อร้าน'].nunique(), '#,##0'),
            ("จำนวนรหัสลูกค้า", df['รหัสลูกค้า_norm'].nunique(), '#,##0')]
    for i,(lbl,val,fmt) in enumerate(kpis):
        W(ws,3,i+1,lbl,bg="2E75B6",fg="FFFFFF",b=True,sz=9)
        W(ws,4,i+1,val,bg="DEEAF1",b=True,sz=13,fg="1F4E79",fmt=fmt)

    if total_target:
        W(ws,3,6,"เป้าหมายรวม (บาท)",bg="C00000",fg="FFFFFF",b=True,sz=9)
        W(ws,4,6,total_target,bg="FCE4D6",b=True,sz=13,fg="C00000",fmt='#,##0')
        ach_bg = "C6EFCE" if overall_ach>=1 else("FFF2CC" if overall_ach>=0.8 else "FFD7D7")
        W(ws,3,7,"% Achievement",bg="C00000",fg="FFFFFF",b=True,sz=9)
        W(ws,4,7,overall_ach,bg=ach_bg,b=True,sz=13,fg="C00000",fmt='0.0%')
        W(ws,3,8,"Gap (จริง - เป้า)",bg="C00000",fg="FFFFFF",b=True,sz=9)
        gap = grand_total - total_target
        W(ws,4,8,gap,bg="E2EFDA" if gap>=0 else "FFD7D7",b=True,sz=12,fg="1F4E79",fmt='+#,##0;-#,##0;0')

    ws.row_dimensions[4].height = 26

    # Monthly table headers — เพิ่ม Target, Gap, %Achieve
    has_target_cols = bool(mt)
    hdrs = ["เดือน","ยอดขาย (บาท)","% ของยอดรวม","จำนวนร้าน","เพิ่ม/ลด vs เดือนก่อน"]
    if has_target_cols:
        hdrs += ["Target (บาท)","Gap","% Achievement"]
    for ci,h in enumerate(hdrs,1):
        W(ws,6,ci,h,bg="1F4E79",fg="FFFFFF",b=True,sz=9,wrap=True)
    ws.row_dimensions[6].height = 28

    # Collect all months (actual + target months)
    all_months = sorted(set(list(months) + list(mt.keys()))) if mt else months

    prev = None
    for ri,m in enumerate(all_months,7):
        mdf = df[df['month']==m]
        sales  = mdf['ราคารวม'].sum()
        pct    = sales/grand_total if grand_total else 0
        chg    = (sales-prev)/prev if prev and m in months else None
        target = mt.get(m)
        alt    = "F2F2F2" if ri%2==0 else None
        in_act = m in months

        W(ws,ri,1,MTH[m],bg=alt,b=True)
        W(ws,ri,2,sales if in_act else "(ยังไม่มีข้อมูล)",bg=alt,fmt='#,##0.00' if in_act else None,
          fg="000000" if in_act else "AAAAAA")
        W(ws,ri,3,pct if in_act else "—",bg=alt,fmt='0.0%' if in_act else None)
        W(ws,ri,4,mdf['ชื่อร้าน'].nunique() if in_act else "—",bg=alt)
        if chg is not None:
            W(ws,ri,5,chg,bg="E2EFDA" if chg>=0 else "FFD7D7",fmt='+0.0%;-0.0%;0.0%')
        else:
            W(ws,ri,5,"—",bg=alt,fg="AAAAAA")

        if has_target_cols:
            W(ws,ri,6,target if target else "—",bg="FCE4D6" if target else alt,
              fmt='#,##0' if target else None,b=bool(target))
            if target and in_act:
                gap = sales-target
                ach = sales/target
                gbg = "E2EFDA" if gap>=0 else "FFD7D7"
                abg = "C6EFCE" if ach>=1 else("FFF2CC" if ach>=0.9 else("FCE4D6" if ach>=0.7 else "FFD7D7"))
                W(ws,ri,7,gap,bg=gbg,fmt='+#,##0;-#,##0;0',sz=9)
                W(ws,ri,8,ach,bg=abg,fmt='0.0%',b=True)
            else:
                for ci in [7,8]: W(ws,ri,ci,"—",bg=alt,fg="AAAAAA")

        if in_act: prev = sales

    tr = 7+len(all_months)
    W(ws,tr,1,"รวม",bg="BDD7EE",b=True)
    W(ws,tr,2,grand_total,bg="2E75B6",fg="FFFFFF",b=True,fmt='#,##0.00')
    W(ws,tr,3,1.0,bg="BDD7EE",b=True,fmt='0.0%')
    if has_target_cols and total_target:
        W(ws,tr,6,total_target,bg="C00000",fg="FFFFFF",b=True,fmt='#,##0')
        gap_all = grand_total-total_target
        W(ws,tr,7,gap_all,bg="E2EFDA" if gap_all>=0 else "FFD7D7",b=True,fmt='+#,##0;-#,##0;0')
        W(ws,tr,8,overall_ach,bg="C6EFCE" if overall_ach>=1 else("FFF2CC" if overall_ach>=0.8 else "FFD7D7"),
          b=True,fmt='0.0%')

    widths = [10,18,12,11,18,15,14,14]
    for i,w_ in enumerate(widths,1): ws.column_dimensions[gc(i)].width=w_

# ════════════════════════════════════════════════════════════════════════════
# SHEET 2: Ranking ร้านค้า
# ════════════════════════════════════════════════════════════════════════════
def build_sheet2(wb, df, grand_total, months):
    ws = wb.create_sheet("2_Ranking_ร้านค้า"); ws.sheet_properties.tabColor = "2E75B6"
    W(ws,1,1,"Ranking ยอดขายรายเดือน – ชื่อร้าน",bg="1F4E79",fg="FFFFFF",b=True,sz=12,h="left",merge_to=4+len(months))

    sm = pivot(df,'ชื่อร้าน',months)
    sm['รวม'] = sm.sum(axis=1); sm['เฉลี่ย'] = sm[months].mean(axis=1)
    sm['%'] = sm['รวม']/grand_total
    sm['_d'] = sm[months].apply(lambda r: sum(1 for i in range(1,len(r)) if r.iloc[i]<r.iloc[i-1] and r.iloc[i-1]>0), axis=1)
    sm = sm.sort_values('รวม',ascending=False)

    hdrs = ['อันดับ','ชื่อร้าน']+[MTH[m] for m in months]+['รวม (บาท)','เฉลี่ย/เดือน','%ยอดรวม','สถานะ']
    for ci,h in enumerate(hdrs,1): W(ws,3,ci,h,bg="1F4E79",fg="FFFFFF",b=True,sz=8,wrap=True)
    ws.row_dimensions[3].height = 30

    for ri_off,(store,row) in enumerate(sm.iterrows(),1):
        ri = ri_off+3; alt = "F0F4FF" if ri_off%2==0 else None
        ws.row_dimensions[ri].height = 15
        W(ws,ri,1,ri_off,bg=alt,b=(ri_off<=10))
        W(ws,ri,2,store,bg=alt,h="left",b=(ri_off<=5),sz=9)
        for ci,m in enumerate(months,3):
            v = row[m]; W(ws,ri,ci,v,bg="FFF0F0" if v==0 else alt,fmt='#,##0',sz=8)
        W(ws,ri,3+len(months),row['รวม'],bg="BDD7EE",b=True,fmt='#,##0.00')
        W(ws,ri,4+len(months),row['เฉลี่ย'],bg=alt,fmt='#,##0',sz=8)
        W(ws,ri,5+len(months),row['%'],bg=alt,fmt='0.00%',sz=8)
        d = row['_d']
        sb = "FFD7D7" if d>=2 else("FFF2CC" if d==1 else "E2EFDA")
        W(ws,ri,6+len(months),"⚠️ เฝ้าระวัง" if d>=2 else("📊 ปกติ" if d==1 else "📈 ดี"),bg=sb,b=True,sz=8)

    tr = len(sm)+4
    W(ws,tr,2,"รวมทั้งหมด",bg="2E75B6",fg="FFFFFF",b=True,h="left")
    for ci,m in enumerate(months,3):
        W(ws,tr,ci,df[df['month']==m]['ราคารวม'].sum(),bg="2E75B6",fg="FFFFFF",b=True,fmt='#,##0')
    W(ws,tr,3+len(months),grand_total,bg="1F4E79",fg="FFFFFF",b=True,fmt='#,##0.00')
    ws.column_dimensions['A'].width=6; ws.column_dimensions['B'].width=30
    for i in range(3,3+len(months)+5): ws.column_dimensions[gc(i)].width=12
    ws.freeze_panes="C4"

# ════════════════════════════════════════════════════════════════════════════
# SHEET 3: Ranking รหัสลูกค้า
# ════════════════════════════════════════════════════════════════════════════
def build_sheet3(wb, df, grand_total, months):
    ws = wb.create_sheet("3_Ranking_รหัสลูกค้า"); ws.sheet_properties.tabColor = "375623"
    W(ws,1,1,"Ranking ยอดขายรายเดือน – รหัสลูกค้า",bg="1F4E79",fg="FFFFFF",b=True,sz=12,h="left",merge_to=5+len(months))

    cm_raw = pivot(df,'รหัสลูกค้า_norm',months)
    cm_raw['รวม'] = cm_raw.sum(axis=1); cm_raw['เฉลี่ย'] = cm_raw[months].mean(axis=1)
    cm_raw['%'] = cm_raw['รวม']/grand_total
    code_store = df.groupby('รหัสลูกค้า_norm')['ชื่อร้าน'].agg(lambda x: x.mode()[0])
    cm_raw['ชื่อร้าน'] = cm_raw.index.map(code_store)
    cm_raw = cm_raw.sort_values('รวม',ascending=False)

    hdrs = ['อันดับ','ชื่อร้าน','รหัสลูกค้า']+[MTH[m] for m in months]+['รวม','เฉลี่ย','%ยอดรวม']
    for ci,h in enumerate(hdrs,1): W(ws,3,ci,h,bg="1F4E79",fg="FFFFFF",b=True,sz=8,wrap=True)
    ws.row_dimensions[3].height = 30

    for ri_off,(code,row) in enumerate(cm_raw.iterrows(),1):
        ri = ri_off+3; alt = "F0F4FF" if ri_off%2==0 else None; ws.row_dimensions[ri].height=14
        W(ws,ri,1,ri_off,bg=alt,b=(ri_off<=10))
        W(ws,ri,2,row.get('ชื่อร้าน',''),bg=alt,h="left",sz=8)
        W(ws,ri,3,code,bg=alt,h="left",sz=8)
        for ci,m in enumerate(months,4):
            v=row[m]; W(ws,ri,ci,v,bg="FFF0F0" if v==0 else alt,fmt='#,##0',sz=8)
        W(ws,ri,4+len(months),row['รวม'],bg="BDD7EE",b=True,fmt='#,##0.00')
        W(ws,ri,5+len(months),row['เฉลี่ย'],bg=alt,fmt='#,##0',sz=8)
        W(ws,ri,6+len(months),row['%'],bg=alt,fmt='0.00%',sz=8)

    tr = len(cm_raw)+4
    W(ws,tr,3,"รวมทั้งหมด",bg="2E75B6",fg="FFFFFF",b=True,h="left")
    for ci,m in enumerate(months,4):
        W(ws,tr,ci,df[df['month']==m]['ราคารวม'].sum(),bg="2E75B6",fg="FFFFFF",b=True,fmt='#,##0')
    W(ws,tr,4+len(months),grand_total,bg="1F4E79",fg="FFFFFF",b=True,fmt='#,##0.00')
    ws.column_dimensions['A'].width=6; ws.column_dimensions['B'].width=26; ws.column_dimensions['C'].width=32
    for i in range(4,4+len(months)+4): ws.column_dimensions[gc(i)].width=12
    ws.freeze_panes="D4"

# ════════════════════════════════════════════════════════════════════════════
# SHEET 4: ยอดขายฟาร์ม
# ════════════════════════════════════════════════════════════════════════════
def build_sheet4(wb, df, grand_total, months):
    ws = wb.create_sheet("4_ยอดขายฟาร์ม"); ws.sheet_properties.tabColor = "375623"
    farm = df[df['รหัสลูกค้า'].str.startswith('ฟาร์ม')].copy()
    ft = farm['ราคารวม'].sum(); nf = farm['รหัสลูกค้า'].nunique()
    W(ws,1,1,"ยอดขายลูกค้าฟาร์ม",bg="375623",fg="FFFFFF",b=True,sz=13,h="left",merge_to=5+len(months))
    W(ws,2,1,f"จำนวน {nf} ฟาร์ม | ยอดรวม {ft:,.2f} บาท | {ft/grand_total:.2%} ของยอดทั้งหมด",
      bg="E2EFDA",fg="375623",b=True,sz=10,h="left",merge_to=5+len(months))
    fm = pivot(farm,'รหัสลูกค้า',months)
    fm['รวม']=fm.sum(axis=1); fm['เฉลี่ย']=fm[months].mean(axis=1); fm['%']=fm['รวม']/grand_total
    fm['เจ้าของ']=fm.index.map(farm.groupby('รหัสลูกค้า')['ชื่อลูกค้า'].first())
    fm=fm.sort_values('รวม',ascending=False)
    for ci,h in enumerate(['อันดับ','ฟาร์ม (รหัสลูกค้า)','เจ้าของ']+[MTH[m] for m in months]+['รวม','เฉลี่ย','%'],1):
        W(ws,4,ci,h,bg="375623",fg="FFFFFF",b=True,sz=8); ws.row_dimensions[4].height=22
    for ri_off,(fc,row) in enumerate(fm.iterrows(),1):
        ri=ri_off+4; alt="F0FFF0" if ri_off%2==0 else None; ws.row_dimensions[ri].height=15
        W(ws,ri,1,ri_off,bg=alt,b=True); W(ws,ri,2,fc,bg=alt,h="left",b=True,sz=9)
        W(ws,ri,3,row.get('เจ้าของ',''),bg=alt,h="left",sz=8)
        for ci,m in enumerate(months,4):
            v=row[m]; W(ws,ri,ci,v,bg="FFF0F0" if v==0 else alt,fmt='#,##0',sz=8)
        W(ws,ri,4+len(months),row['รวม'],bg="A9D18E",b=True,fmt='#,##0.00')
        W(ws,ri,5+len(months),row['เฉลี่ย'],bg=alt,fmt='#,##0',sz=8)
        W(ws,ri,6+len(months),row['%'],bg=alt,fmt='0.00%',sz=8)
    ws.column_dimensions['A'].width=6; ws.column_dimensions['B'].width=34; ws.column_dimensions['C'].width=26
    for i in range(4,4+len(months)+4): ws.column_dimensions[gc(i)].width=12

# ════════════════════════════════════════════════════════════════════════════
# SHEET 5: เป้าหมาย PLC & Recco
# ════════════════════════════════════════════════════════════════════════════
def build_sheet5(wb, df, grand_total, months, plc_target, plc_deadline, recco_target, recco_deadline):
    ws = wb.create_sheet("5_เป้าหมาย_PLC_Recco"); ws.sheet_properties.tabColor = "C00000"
    W(ws,1,1,"ติดตามเป้าหมายยอดขาย",bg="C00000",fg="FFFFFF",b=True,sz=13,h="left",merge_to=12)

    plc_names = ['บริษัท วีว่าพรีเมี่ยม เพ็ท สโตร์ จำกัด','บริษัท วีว่า เพ็ทสโตร์ จำกัด','บริษัท วีว่า เพ็ท สโตร์ จำกัด']
    plc_df    = df[df['ชื่อลูกค้า'].isin(plc_names)]
    plc_act   = plc_df['ราคารวม'].sum()
    recco_df  = df[df['ชื่อลูกค้า']=='บริษัท เรคโค เพ็ท จำกัด']
    recco_act = recco_df['ราคารวม'].sum()

    def block(sc, name, target, actual, deadline, store_df):
        W(ws,3,sc,name,bg="2E75B6",fg="FFFFFF",b=True,sz=11,merge_to=sc+5)
        rows = [("เป้าหมาย",target,'#,##0'),("ยอดปัจจุบัน",actual,'#,##0.00'),
                ("ยอดที่เหลือ",max(0,target-actual),'#,##0.00'),
                ("% บรรลุเป้า",actual/target if target else 0,'0.0%'),
                ("กำหนดเวลา",deadline,None)]
        for ri,(lbl,val,fmt) in enumerate(rows,4):
            W(ws,ri,sc,lbl,bg="DEEAF1",b=True,sz=9,merge_to=sc+1)
            pct = actual/target if target else 0
            vbg = "C6EFCE" if pct>=1 else("FFF2CC" if pct>=0.7 else "FFD7D7") if lbl=="% บรรลุเป้า" else "FFFFFF"
            W(ws,ri,sc+2,val,bg=vbg,b=(lbl=="% บรรลุเป้า"),sz=11,fmt=fmt,merge_to=sc+5)
        W(ws,9,sc,"เดือน",bg="2E75B6",fg="FFFFFF",b=True,sz=8)
        W(ws,9,sc+1,"ยอดขาย (บาท)",bg="2E75B6",fg="FFFFFF",b=True,sz=8)
        W(ws,9,sc+2,"% ของเป้า",bg="2E75B6",fg="FFFFFF",b=True,sz=8)
        for ri,m in enumerate(months,10):
            mv = store_df[store_df['month']==m]['ราคารวม'].sum()
            alt = "F0F4FF" if ri%2==0 else None
            W(ws,ri,sc,MTH[m],bg=alt,sz=8)
            W(ws,ri,sc+1,mv,bg=alt,fmt='#,##0.00',sz=8)
            W(ws,ri,sc+2,mv/target if target else 0,bg=alt,fmt='0.0%',sz=8)
        sr = 10+len(months)
        W(ws,sr,sc,"สะสม",bg="BDD7EE",b=True,sz=9)
        W(ws,sr,sc+1,actual,bg="BDD7EE",b=True,fmt='#,##0.00')
        W(ws,sr,sc+2,actual/target if target else 0,bg="BDD7EE",b=True,fmt='0.0%')

    block(1,"Pet Lover Centre",plc_target,plc_act,plc_deadline,plc_df)
    block(8,"บริษัท เรคโค เพ็ท จำกัด",recco_target,recco_act,recco_deadline,recco_df)
    for i in range(1,15): ws.column_dimensions[gc(i)].width=14

# ════════════════════════════════════════════════════════════════════════════
# SHEET 6: ภาพรวมสินค้า  ← เพิ่ม จำนวนชิ้น
# ════════════════════════════════════════════════════════════════════════════
def build_sheet6(wb, df, grand_total):
    ws = wb.create_sheet("6_ภาพรวมสินค้า"); ws.sheet_properties.tabColor = "7030A0"
    W(ws,1,1,"ภาพรวมยอดขายสินค้า – แยกหมวดหมู่",bg="3B1F78",fg="FFFFFF",b=True,sz=12,h="left",merge_to=5)
    cats = df.groupby('หมวดหมู่').agg(
        ยอดขาย=('ราคารวม','sum'), จำนวนชิ้น=('จำนวน','sum'),
        จำนวน_SKU=('prod_grp','nunique')).reset_index().sort_values('ยอดขาย',ascending=False)
    cats['%'] = cats['ยอดขาย']/grand_total
    for ci,h in enumerate(["หมวดหมู่","ยอดขาย (บาท)","% ของยอดรวม","จำนวนชิ้น","จำนวน SKU"],1):
        W(ws,3,ci,h,bg="3B1F78",fg="FFFFFF",b=True,sz=9); ws.row_dimensions[3].height=22
    for ri_off,row in enumerate(cats.itertuples(),1):
        ri=ri_off+3; cbg=CAT_CLR.get(row.หมวดหมู่,'F2F2F2')
        alt = "F5F0FF" if ri_off%2==0 else None
        W(ws,ri,1,row.หมวดหมู่,bg=cbg,h="left",b=True,sz=9)
        W(ws,ri,2,row.ยอดขาย,bg=alt,fmt='#,##0.00')
        W(ws,ri,3,row._5,bg=alt,fmt='0.00%')
        W(ws,ri,4,int(row[3]),bg="D9F0E0",fmt='#,##0',b=True)   # ← ชิ้น (col index 3)
        W(ws,ri,5,int(row[4]),bg=alt)
    tr=len(cats)+4
    W(ws,tr,1,"รวม",bg="BDD7EE",b=True,h="left")
    W(ws,tr,2,grand_total,bg="2E75B6",fg="FFFFFF",b=True,fmt='#,##0.00')
    W(ws,tr,3,1.0,bg="BDD7EE",b=True,fmt='0.0%')
    W(ws,tr,4,int(df['จำนวน'].sum()),bg="375623",fg="FFFFFF",b=True,fmt='#,##0')
    for i,w_ in enumerate([18,16,14,13,11],1): ws.column_dimensions[gc(i)].width=w_

# ════════════════════════════════════════════════════════════════════════════
# SHEET 7: Ranking สินค้า  ← เพิ่ม จำนวนชิ้น
# ════════════════════════════════════════════════════════════════════════════
def build_sheet7(wb, df, grand_total, months, prod_names, prod_cat):
    ws = wb.create_sheet("7_Ranking_สินค้า"); ws.sheet_properties.tabColor = "7030A0"
    W(ws,1,1,"Ranking ยอดขายสินค้า (กลุ่มรหัส 6 ตัว)",bg="3B1F78",fg="FFFFFF",b=True,sz=12,h="left",merge_to=6+len(months)+3)
    pm = pivot(df,'prod_grp',months,'ราคารวม')
    qm = pivot(df,'prod_grp',months,'จำนวน')       # ← pivot ชิ้น
    pm['รวมบาท']=pm.sum(axis=1); pm['รวมชิ้น']=qm.sum(axis=1)
    pm['เฉลี่ยบาท']=pm[months].mean(axis=1); pm['เฉลี่ยชิ้น']=qm[months].mean(axis=1)
    pm['%']=pm['รวมบาท']/grand_total
    pm['ชื่อ']=pm.index.map(prod_names); pm['หมวด']=pm.index.map(prod_cat)
    pm=pm.sort_values('รวมบาท',ascending=False)
    hdrs=['อันดับ','รหัส','ชื่อสินค้า','หมวดหมู่']+[MTH[m] for m in months]+\
         ['รวม (บาท)','รวม (ชิ้น)','เฉลี่ย/เดือน (บาท)','เฉลี่ย/เดือน (ชิ้น)','%ยอดรวม']
    for ci,h in enumerate(hdrs,1): W(ws,3,ci,h,bg="3B1F78",fg="FFFFFF",b=True,sz=8,wrap=True)
    ws.row_dimensions[3].height=36
    for ri_off,(code,row) in enumerate(pm.iterrows(),1):
        ri=ri_off+3; alt="F5F0FF" if ri_off%2==0 else None; ws.row_dimensions[ri].height=14
        W(ws,ri,1,ri_off,bg=alt,b=(ri_off<=10)); W(ws,ri,2,code,bg=alt,sz=8)
        W(ws,ri,3,row['ชื่อ'],bg=alt,h="left",sz=8,wrap=True)
        W(ws,ri,4,row['หมวด'],bg=CAT_CLR.get(row['หมวด'],'F2F2F2'),sz=8)
        for ci,m in enumerate(months,5):
            v=row[m]; W(ws,ri,ci,v,bg="FFF0F0" if v==0 else alt,fmt='#,##0',sz=8)
        cn=5+len(months)
        W(ws,ri,cn,  row['รวมบาท'],  bg="E8D5F5",b=True,fmt='#,##0.00'); cn+=1
        W(ws,ri,cn,  row['รวมชิ้น'], bg="D9F0E0",b=True,fmt='#,##0');    cn+=1  # ← ชิ้น
        W(ws,ri,cn,  row['เฉลี่ยบาท'],bg=alt,fmt='#,##0',sz=8);          cn+=1
        W(ws,ri,cn,  row['เฉลี่ยชิ้น'],bg=alt,fmt='#,##0',sz=8);         cn+=1  # ← ชิ้น
        W(ws,ri,cn,  row['%'],        bg=alt,fmt='0.00%',sz=8)
    ws.column_dimensions['A'].width=6; ws.column_dimensions['B'].width=9
    ws.column_dimensions['C'].width=34; ws.column_dimensions['D'].width=13
    for i in range(5,5+len(months)+6): ws.column_dimensions[gc(i)].width=11
    ws.freeze_panes="E4"

# ════════════════════════════════════════════════════════════════════════════
# SHEET 8 & 9: สินค้า × Top10  ← เพิ่ม จำนวนชิ้น
# ════════════════════════════════════════════════════════════════════════════
def build_sheets_89(wb, df, grand_total, months, prod_names, prod_cat):
    # layout per product row:
    # cols 1-3: รหัส, ชื่อ, หมวด
    # cols 4..3+M: ยอด (บาท) รายเดือน
    # cols 4+M..3+2M: จำนวน (ชิ้น) รายเดือน  ← ใหม่
    # cols 4+2M: รวม (บาท), 5+2M: รวม (ชิ้น), 6+2M: %
    M = len(months)
    total_cols = 3 + M + M + 3   # รหัส+ชื่อ+หมวด + บาทxM + ชิ้นxM + รวมบาท+รวมชิ้น+%

    for sheet_num,group_col,sname,tab,hbg in [
        (8,'ชื่อร้าน',       '8_สินค้า×Top10ร้าน','ED7D31','ED7D31'),
        (9,'รหัสลูกค้า_norm','9_สินค้า×Top10รหัส','FF0000','C00000'),
    ]:
        ws=wb.create_sheet(sname); ws.sheet_properties.tabColor=tab
        W(ws,1,1,f"Ranking สินค้า – Top10 {group_col}  |  ยอด (บาท) + จำนวน (ชิ้น) รายเดือน",
          bg="1F4E79",fg="FFFFFF",b=True,sz=12,h="left",merge_to=total_cols)

        top10=(df.groupby(group_col)['ราคารวม'].sum()
                 .sort_values(ascending=False).head(10).index.tolist())
        ri=3
        for rank_i,grp_val in enumerate(top10,1):
            grp_total=df[df[group_col]==grp_val]['ราคารวม'].sum()
            grp_qty  =df[df[group_col]==grp_val]['จำนวน'].sum()

            # ── Store/code header ─────────────────────────────────────────
            W(ws,ri,1,f"#{rank_i} {grp_val}",
              bg=hbg,fg="FFFFFF",b=True,sz=10,h="left",merge_to=total_cols)
            ri+=1

            # ── Column headers ────────────────────────────────────────────
            hbg2="FCE4D6" if sheet_num==8 else "FCE9E9"
            hfg2="843C0C" if sheet_num==8 else "C00000"

            W(ws,ri,1,"รหัส",bg=hbg2,fg=hfg2,b=True,sz=8)
            W(ws,ri,2,"ชื่อสินค้า",bg=hbg2,fg=hfg2,b=True,sz=8)
            W(ws,ri,3,"หมวดหมู่",bg=hbg2,fg=hfg2,b=True,sz=8)
            # บาทรายเดือน
            for ci,m in enumerate(months,4):
                W(ws,ri,ci,f"{MTH[m]}\n(บาท)",bg="BDD7EE",fg="1F4E79",b=True,sz=8,wrap=True)
            # ชิ้นรายเดือน
            for ci,m in enumerate(months,4+M):
                W(ws,ri,ci,f"{MTH[m]}\n(ชิ้น)",bg="D9F0E0",fg="375623",b=True,sz=8,wrap=True)
            # summary
            W(ws,ri,4+2*M,"รวม\n(บาท)",  bg="BDD7EE",fg="1F4E79",b=True,sz=8,wrap=True)
            W(ws,ri,5+2*M,"รวม\n(ชิ้น)", bg="D9F0E0",fg="375623",b=True,sz=8,wrap=True)
            W(ws,ri,6+2*M,"% ของกลุ่ม",  bg=hbg2,fg=hfg2,b=True,sz=8,wrap=True)
            ws.row_dimensions[ri].height=30
            ri+=1

            # ── Data rows ─────────────────────────────────────────────────
            sub=df[df[group_col]==grp_val]
            pm=pivot(sub,'prod_grp',months,'ราคารวม')
            qm=pivot(sub,'prod_grp',months,'จำนวน')
            pm['รวมบาท']=pm.sum(axis=1); pm['รวมชิ้น']=qm.sum(axis=1)
            pm=pm.sort_values('รวมบาท',ascending=False)

            for ri_off2,(pcode,prow) in enumerate(pm.iterrows(),1):
                alt2="F9F5FF" if ri_off2%2==0 else None
                ws.row_dimensions[ri].height=14
                W(ws,ri,1,pcode,bg=alt2,sz=8)
                W(ws,ri,2,prod_names.get(pcode,pcode),bg=alt2,h="left",sz=8,wrap=False)
                W(ws,ri,3,prod_cat.get(pcode,''),bg=alt2,sz=8)
                # บาทรายเดือน
                for ci,m in enumerate(months,4):
                    v=prow.get(m,0)
                    W(ws,ri,ci,v,bg="FFF0F0" if v==0 else alt2,fmt='#,##0',sz=8)
                # ชิ้นรายเดือน (from qm)
                qrow=qm.loc[pcode] if pcode in qm.index else pd.Series({m:0 for m in months})
                for ci,m in enumerate(months,4+M):
                    qv=qrow.get(m,0)
                    W(ws,ri,ci,int(qv),bg="F0FFF0" if qv>0 else "FFF0F0",fmt='#,##0',sz=8)
                # summary
                W(ws,ri,4+2*M,prow['รวมบาท'],bg="BDD7EE",b=True,fmt='#,##0')
                W(ws,ri,5+2*M,int(prow['รวมชิ้น']),bg="D9F0E0",b=True,fmt='#,##0')
                W(ws,ri,6+2*M,prow['รวมบาท']/grp_total if grp_total else 0,
                  bg=alt2,fmt='0.0%',sz=8)
                ri+=1

            # ── Group total row ───────────────────────────────────────────
            W(ws,ri,1,"รวม",bg=hbg,fg="FFFFFF",b=True)
            ws.merge_cells(start_row=ri,start_column=1,end_row=ri,end_column=3)
            for ci,m in enumerate(months,4):
                v=sub[sub['month']==m]['ราคารวม'].sum()
                W(ws,ri,ci,v,bg="2E75B6",fg="FFFFFF",b=True,fmt='#,##0')
            for ci,m in enumerate(months,4+M):
                qv=sub[sub['month']==m]['จำนวน'].sum()
                W(ws,ri,ci,int(qv),bg="375623",fg="FFFFFF",b=True,fmt='#,##0')
            W(ws,ri,4+2*M,grp_total,bg="1F4E79",fg="FFFFFF",b=True,fmt='#,##0')
            W(ws,ri,5+2*M,int(grp_qty),bg="1E5C1E",fg="FFFFFF",b=True,fmt='#,##0')
            W(ws,ri,6+2*M,1.0,bg=hbg,fg="FFFFFF",b=True,fmt='0.0%')
            ri+=2   # gap

        # ── Column widths ─────────────────────────────────────────────────
        ws.column_dimensions['A'].width=10
        ws.column_dimensions['B'].width=32
        ws.column_dimensions['C'].width=13
        for i in range(4,4+M):   ws.column_dimensions[gc(i)].width=11   # บาท
        for i in range(4+M,4+2*M): ws.column_dimensions[gc(i)].width=9  # ชิ้น
        ws.column_dimensions[gc(4+2*M)].width=12
        ws.column_dimensions[gc(5+2*M)].width=10
        ws.column_dimensions[gc(6+2*M)].width=10
        ws.freeze_panes="D2"

# ════════════════════════════════════════════════════════════════════════════
# SHEET 10: Forecast ร้านค้า
# ════════════════════════════════════════════════════════════════════════════
def build_sheet10(wb, df, grand_total, months):
    ACT=months; FCI=list(range(max(months)+1,13))
    if not FCI: FCI=list(range(6,13))
    ws=wb.create_sheet("10_Forecast_ร้านค้า"); ws.sheet_properties.tabColor="1F4E79"
    W(ws,1,1,f"📊 Forecast & วิเคราะห์ร้านค้า  |  ยอดจริง {MTH[min(ACT)]}–{MTH[max(ACT)]} + คาดการณ์ {MTH[min(FCI)]}–{MTH[max(FCI)]}",
      bg="1F4E79",fg="FFFFFF",b=True,sz=12,h="left",merge_to=3+len(ACT)+1+len(FCI)*2+2+3)
    for ci,(cat,st) in enumerate(STORE_CAT.items(),1):
        W(ws,2,ci,cat,bg=st['bg'],fg=st['fg'],b=True,sz=8)
    ws.row_dimensions[2].height=16
    sm=pivot(df,'ชื่อร้าน',ACT)
    rows=[]
    for store,row in sm.iterrows():
        v=row.values.astype(float); sl,r2,fc,wma=make_forecast(v,ACT,FCI)
        fc_ma={m:wma for m in FCI}; active=sum(1 for x in v if x>0)
        peak_m=ACT[int(np.argmax(v))]; peak_v=float(np.max(v)); cur=float(v[-1])
        r2a=np.mean(v[len(v)//2:]); r2b=max(np.mean(v[:len(v)//2]),1)
        if   sl>8000 and r2>0.5:                cat='🌟 STAR'
        elif sl>2000 and r2>0.3:                cat='📈 GROWTH'
        elif r2a>r2b*1.2 and active>=3:         cat='🔄 RECOVERING'
        elif peak_v>0 and cur<peak_v*.55 and active>=3: cat='📉 DECLINING'
        elif sl<-3000 and active>=3:            cat='⚠️ WARNING'
        elif active<=2:                         cat='⚡ INACTIVE'
        elif active==3:                         cat='🔵 SPORADIC'
        else:                                   cat='😴 STABLE'
        action_m={'🌟 STAR':"ขยาย SKU / volume discount",'📈 GROWTH':"ติดตาม / เสนอ bundle",
                  '🔄 RECOVERING':"โปร + ติดตาม 2 เดือน",'📉 DECLINING':f"ยอดตกจาก peak {MTH[peak_m]} → โทรหา",
                  '⚠️ WARNING':"เร่งด่วน! ตรวจสอบสาเหตุ",'⚡ INACTIVE':f"ไม่มียอดใน M{max(ACT)} → follow up",
                  '🔵 SPORADIC':"ซื้อไม่สม่ำเสมอ → สร้าง habit",'😴 STABLE':"รักษาฐาน / เสนอสินค้าใหม่"}
        rows.append({'ชื่อร้าน':store,'cat':cat,'action':action_m.get(cat,''),
            **{f'M{m}':v[i] for i,m in enumerate(ACT)},
            'total':v.sum(),'pct':v.sum()/grand_total,'active':active,
            'peak_m':peak_m,'peak_v':peak_v,'cur':cur,
            'vs_peak':(cur-peak_v)/peak_v if peak_v>0 else 0,'slope':sl,'r2':r2,'wma':wma,
            'fc_lin':fc,'fc_ma':fc_ma,'fc_lin_sum':sum(fc.values()),'fc_ma_sum':sum(fc_ma.values())})
    CAT_ORD=['🌟 STAR','📈 GROWTH','🔄 RECOVERING','📉 DECLINING','⚠️ WARNING','🔵 SPORADIC','😴 STABLE','⚡ INACTIVE']
    stores=pd.DataFrame(rows); stores['_so']=stores['cat'].map({v:i for i,v in enumerate(CAT_ORD)})
    stores=stores.sort_values(['_so','total'],ascending=[True,False])
    W(ws,3,1,"📌",bg="2E75B6",fg="FFFFFF",b=True)
    ci2=2
    for cat in CAT_ORD:
        sub=stores[stores['cat']==cat]
        if len(sub)>0:
            st=STORE_CAT.get(cat,{'bg':'F2F2F2','fg':'000000'})
            W(ws,3,ci2,f"{cat}: {len(sub)} ร้าน",bg=st['bg'],fg=st['fg'],b=True,sz=8,h="left"); ci2+=1
    ws.row_dimensions[3].height=16; ws.row_dimensions[4].height=36
    HDRS=[('#','1F4E79'),('ชื่อร้าน','1F4E79'),('สถานะ','1F4E79')]
    for m in ACT: HDRS.append((f"{MTH[m]}\nจริง","2E75B6"))
    HDRS+=[('รวมจริง','1F4E79'),('Peak\nเดือน','595959'),('Peak\nมูลค่า','595959'),('vs Peak','595959')]
    for m in FCI: HDRS.append((f"🔵{MTH[m]}\nLinear","375623"))
    HDRS.append(('FC รวม\nLinear','1E5C1E'))
    for m in FCI: HDRS.append((f"🟠{MTH[m]}\nWMA","C55A11"))
    HDRS+=[('FC รวม\nWMA','7B2D00'),('slope/\nเดือน','595959'),('%ยอดรวม','595959'),('📋 แนะนำ','1F4E79')]
    for ci,(h,bg) in enumerate(HDRS,1): W(ws,4,ci,h,bg=bg,fg="FFFFFF",b=True,sz=8,wrap=True)
    for ri_off,(_,row) in enumerate(stores.iterrows(),1):
        r=ri_off+4; ws.row_dimensions[r].height=15
        st=STORE_CAT.get(row['cat'],{'bg':None,'fg':'000000'}); alt="F7FBFF" if ri_off%2==0 else None; ci=1
        W(ws,r,ci,ri_off,bg=alt,b=(ri_off<=10)); ci+=1
        W(ws,r,ci,row['ชื่อร้าน'],bg=alt,h="left",b=(row['cat'] in ['🌟 STAR','📈 GROWTH']),sz=9); ci+=1
        W(ws,r,ci,row['cat'],bg=st['bg'],fg=st['fg'],b=True,sz=8); ci+=1
        for m in ACT:
            v=row[f'M{m}']; mbg="FFF0A0" if m==row['peak_m'] and v>0 else("FFF0F0" if v==0 else alt)
            W(ws,r,ci,v,bg=mbg,fmt='#,##0',sz=8); ci+=1
        W(ws,r,ci,row['total'],bg="BDD7EE",b=True,fmt='#,##0.00'); ci+=1
        W(ws,r,ci,MTH.get(row['peak_m'],''),bg=alt,sz=8); ci+=1
        W(ws,r,ci,row['peak_v'],bg=alt,fmt='#,##0',sz=8); ci+=1
        vp=row['vs_peak']; vbg="FFD7D7" if vp<-0.4 else("FFF2CC" if vp<-0.15 else("E2EFDA" if vp>=0 else alt))
        W(ws,r,ci,vp,bg=vbg,fmt='+0%;-0%;0%',b=(abs(vp)>0.3),sz=8); ci+=1
        for m in FCI: W(ws,r,ci,row['fc_lin'][m],bg="E2EFDA" if row['fc_lin'][m]>0 else "FFF0F0",fmt='#,##0',sz=8,italic=True); ci+=1
        W(ws,r,ci,row['fc_lin_sum'],bg="A9D18E",b=True,fmt='#,##0.00'); ci+=1
        for m in FCI: W(ws,r,ci,row['fc_ma'][m],bg="FFF2E3",fmt='#,##0',sz=8,italic=True); ci+=1
        W(ws,r,ci,row['fc_ma_sum'],bg="F4B183",b=True,fmt='#,##0.00'); ci+=1
        W(ws,r,ci,row['slope'],bg="E2EFDA" if row['slope']>0 else "FFD7D7",fmt='+#,##0;-#,##0;0',sz=8); ci+=1
        W(ws,r,ci,row['pct'],bg=alt,fmt='0.00%',sz=8); ci+=1
        W(ws,r,ci,row['action'],bg=alt,h="left",sz=8,wrap=True)
    ws.column_dimensions['A'].width=5; ws.column_dimensions['B'].width=30; ws.column_dimensions['C'].width=18
    for i in range(4,4+len(ACT)+4): ws.column_dimensions[gc(i)].width=11
    for i in range(4+len(ACT)+4,4+len(ACT)+4+len(FCI)*2+2): ws.column_dimensions[gc(i)].width=9
    lc=4+len(ACT)+4+len(FCI)*2+2
    ws.column_dimensions[gc(lc)].width=9; ws.column_dimensions[gc(lc+1)].width=8; ws.column_dimensions[gc(lc+2)].width=30
    ws.freeze_panes=f"{gc(4+len(ACT))}5"

# ════════════════════════════════════════════════════════════════════════════
# SHEET 11: Forecast สินค้า × ร้านค้า  ← เพิ่ม จำนวนชิ้น
# ════════════════════════════════════════════════════════════════════════════
def build_sheet11(wb, df, grand_total, months, prod_names, prod_cat):
    ws=wb.create_sheet("11_Forecast_สินค้า"); ws.sheet_properties.tabColor="7030A0"
    W(ws,1,1,"📦 วิเคราะห์สินค้า × ร้านค้า  |  กำลังโต / เคยดีแล้วลด / หยุดซื้อ",
      bg="3B1F78",fg="FFFFFF",b=True,sz=12,h="left",merge_to=19)
    for ci,(k,(bg,fg,lbl)) in enumerate(PAIR_STYLE.items(),1):
        W(ws,2,ci,lbl,bg=bg,fg=fg,sz=8,b=True)
    ws.row_dimensions[2].height=15
    W(ws,3,1,"💡 header row = ยอดรวมสินค้าทุกร้าน  |  แถวย่อย = แต่ละร้านที่ซื้อสินค้านี้",
      bg="F3EEFF",fg="3B1F78",sz=8,h="left",merge_to=19)
    ws.row_dimensions[3].height=14
    P_HDRS=[('#','3B1F78'),('รหัส','3B1F78'),('ชื่อสินค้า','3B1F78'),('หมวดหมู่','3B1F78'),
            ('ชื่อร้าน','3B1F78'),('สถานะ','3B1F78')]
    for m in months: P_HDRS.append((f"{MTH[m]}\nจริง","2E75B6"))
    P_HDRS+=[('รวม (บาท)','1F4E79'),('รวม (ชิ้น)','375623'),  # ← เพิ่มชิ้น
             ('%สินค้า','595959'),('Peak','595959'),
             ('Early\nAvg','595959'),('Late\nAvg','595959'),('Trend\n%','595959'),('💬 Insight','3B1F78')]
    for ci,(h,bg) in enumerate(P_HDRS,1): W(ws,4,ci,h,bg=bg,fg="FFFFFF",b=True,sz=8,wrap=True)
    ws.row_dimensions[4].height=36

    cross=df.groupby(['prod_grp','ชื่อร้าน','month'])['ราคารวม'].sum().unstack(fill_value=0)
    cross_q=df.groupby(['prod_grp','ชื่อร้าน','month'])['จำนวน'].sum().unstack(fill_value=0)
    for m in months:
        if m not in cross.columns:   cross[m]=0
        if m not in cross_q.columns: cross_q[m]=0
    cross=cross[months]; cross_q=cross_q[months]
    cross['total']=cross.sum(axis=1); cross['total_q']=cross_q.sum(axis=1)
    cross['early_avg']=cross[months[:3]].mean(axis=1); cross['late_avg']=cross[months[-2:]].mean(axis=1)
    cross['trend_pct']=(cross['late_avg']-cross['early_avg'])/cross['early_avg'].replace(0,np.nan)
    cross['active']=(cross[months]>0).sum(axis=1)
    cross['peak_m']=cross[months].idxmax(axis=1); cross['peak_v']=cross[months].max(axis=1)
    cross['last_v']=cross[months[-1]]
    def cpair(row):
        tp=row['trend_pct']; ea=row['early_avg']; la=row['late_avg']
        if row['last_v']==0 and ea>5000: return 'STOPPED'
        if tp<-0.8 and ea>5000:          return 'DROPPED'
        if la>3000 and tp>0.3:            return 'GROWING'
        if row['active']<=1:              return 'ONE_TIME'
        if row['active']==2:              return 'SPORADIC'
        if tp<-0.2:                       return 'DECLINING'
        if tp>0.1:                        return 'GROWING'
        return 'STABLE'
    cross['pair_status']=cross.apply(cpair,axis=1)
    cross=cross.reset_index(); cross.columns.name=None
    cross=cross[cross['total']>1000].copy()

    prod_tot  =df.groupby(['prod_grp','month'])['ราคารวม'].sum().unstack(fill_value=0)
    prod_tot_q=df.groupby(['prod_grp','month'])['จำนวน'].sum().unstack(fill_value=0)
    for m in months:
        if m not in prod_tot.columns:   prod_tot[m]=0
        if m not in prod_tot_q.columns: prod_tot_q[m]=0
    prod_tot=prod_tot[months]; prod_tot_q=prod_tot_q[months]
    prod_tot['total']=prod_tot.sum(axis=1); prod_tot['total_q']=prod_tot_q.sum(axis=1)
    prod_tot['early']=prod_tot[months[:3]].mean(axis=1); prod_tot['late']=prod_tot[months[-2:]].mean(axis=1)
    prod_tot['trend']=(prod_tot['late']-prod_tot['early'])/prod_tot['early'].replace(0,np.nan)
    prod_tot=prod_tot.sort_values('total',ascending=False)
    SORT_PS={'GROWING':0,'STABLE':1,'DECLINING':2,'DROPPED':3,'STOPPED':4,'SPORADIC':5,'ONE_TIME':6}

    ri=5; rank=0
    for pcode in prod_tot.index:
        if pcode not in prod_names: continue
        pname=prod_names[pcode]; pcat=prod_cat.get(pcode,'')
        prows=cross[cross['prod_grp']==pcode].copy()
        if len(prows)==0: continue
        ptotal=prod_tot.loc[pcode,'total']; ptotal_q=prod_tot.loc[pcode,'total_q']
        pearly=prod_tot.loc[pcode,'early']; plate=prod_tot.loc[pcode,'late']
        ptrend=prod_tot.loc[pcode,'trend']
        rank+=1
        pbg='E2EFDA' if ptrend>0.2 else('FCE4D6' if ptrend<-0.3 else 'DEEAF1')
        ptlbl='📈 GROWTH' if ptrend>0.2 else('📉 DECLINING' if ptrend<-0.3 else '➡️ STABLE')
        ws.row_dimensions[ri].height=17
        W(ws,ri,1,rank,bg=pbg,b=True,sz=9); W(ws,ri,2,pcode,bg=pbg,b=True,sz=9)
        W(ws,ri,3,pname,bg=pbg,b=True,sz=9,h="left",wrap=True)
        W(ws,ri,4,pcat,bg=CAT_CLR.get(pcat,'F2F2F2'),sz=8,b=True)
        W(ws,ri,5,f"รวม {len(prows)} ร้าน",bg=pbg,sz=8,italic=True)
        W(ws,ri,6,ptlbl,bg=pbg,b=True,sz=8)
        for ci2,m in enumerate(months,7):
            v=prod_tot.loc[pcode,m] if m in prod_tot.columns else 0
            W(ws,ri,ci2,v,bg=pbg,b=True,fmt='#,##0',sz=9)
        cn=7+len(months)
        W(ws,ri,cn,ptotal,  bg=pbg,b=True,fmt='#,##0.00',sz=9);    cn+=1
        W(ws,ri,cn,ptotal_q,bg="D9F0E0",b=True,fmt='#,##0',sz=9);  cn+=1  # ← ชิ้น
        W(ws,ri,cn,ptotal/grand_total,bg=pbg,fmt='0.00%',sz=8);     cn+=1
        peak_m2=prod_tot.loc[pcode,months].idxmax()
        W(ws,ri,cn,MTH.get(peak_m2,''),bg=pbg,sz=8);                cn+=1
        W(ws,ri,cn,pearly,bg=pbg,fmt='#,##0',sz=8);                 cn+=1
        W(ws,ri,cn,plate,bg=pbg,fmt='#,##0',sz=8);                  cn+=1
        tbg="E2EFDA" if ptrend>0 else "FCE4D6"
        W(ws,ri,cn,ptrend if not np.isnan(ptrend) else 0,bg=tbg,fmt='+0%;-0%;0%',b=True,sz=8); cn+=1
        ng=len(prows[prows['pair_status']=='GROWING']); ns=len(prows[prows['pair_status'].isin(['STOPPED','DROPPED'])])
        W(ws,ri,cn,f"{ng} ร้านกำลังโต"+(f"  |  {ns} ร้านหยุดซื้อ" if ns>0 else ""),bg=pbg,h="left",sz=8)
        ri+=1

        prows=prows.copy(); prows['_ps']=prows['pair_status'].map(SORT_PS).fillna(9)
        prows=prows.sort_values(['_ps','total'],ascending=[True,False])
        for _,pr in prows.iterrows():
            ws.row_dimensions[ri].height=14; ps=pr['pair_status']
            pb2,pf2,plbl2=PAIR_STYLE.get(ps,('F2F2F2','000000','')); alt2="FDFBFF" if ri%2==0 else None
            for ci2 in [1,2,3,4]: W(ws,ri,ci2,'',bg=alt2)
            W(ws,ri,5,pr['ชื่อร้าน'],bg=alt2,h="left",sz=8)
            W(ws,ri,6,plbl2,bg=pb2,fg=pf2,b=True,sz=8)
            for ci2,m in enumerate(months,7):
                v=pr.get(m,0); mbg="FFF0A0" if m==pr['peak_m'] and v>0 else("FFF0F0" if v==0 else alt2)
                W(ws,ri,ci2,v,bg=mbg,fmt='#,##0',sz=8)
            cn=7+len(months)
            W(ws,ri,cn,pr['total'],   bg=alt2,b=(ps=='GROWING'),fmt='#,##0',sz=8);           cn+=1
            W(ws,ri,cn,pr['total_q'], bg="F0FFF0" if pr['total_q']>0 else "FFF0F0",fmt='#,##0',sz=8); cn+=1  # ← ชิ้น
            W(ws,ri,cn,pr['total']/ptotal if ptotal>0 else 0,bg=alt2,fmt='0.0%',sz=8);       cn+=1
            W(ws,ri,cn,MTH.get(pr['peak_m'],''),bg=alt2,sz=8);                               cn+=1
            W(ws,ri,cn,pr['early_avg'],bg=alt2,fmt='#,##0',sz=8);                            cn+=1
            W(ws,ri,cn,pr['late_avg'],bg="FFF0F0" if pr['late_avg']==0 else alt2,fmt='#,##0',sz=8); cn+=1
            tp2=pr['trend_pct'] if not(isinstance(pr['trend_pct'],float) and np.isnan(pr['trend_pct'])) else 0
            W(ws,ri,cn,tp2,bg="E2EFDA" if tp2>0.1 else("FFD7D7" if tp2<-0.3 else alt2),fmt='+0%;-0%;0%',sz=8); cn+=1
            if ps=='GROWING':    ins=f"↑ avg {pr['late_avg']:,.0f}฿ → ดัน SKU"
            elif ps=='STOPPED':  ins=f"⛔ เคย avg {pr['early_avg']:,.0f}฿ → โทรหาด่วน"
            elif ps=='DROPPED':  ins=f"⚠️ ลดจาก {pr['peak_v']:,.0f}฿ → ตรวจสอบ"
            elif ps=='DECLINING': ins=f"↓ {tp2:.0%} → เสนอโปรฯ"
            elif ps=='SPORADIC':  ins="ซื้อสลับ → สร้างความสม่ำเสมอ"
            elif ps=='ONE_TIME':  ins="ซื้อครั้งเดียว → follow up"
            else:                 ins="คงที่ → รักษาฐาน"
            W(ws,ri,cn,ins,bg=alt2,h="left",sz=8,wrap=True); ri+=1
        ri+=1

    ws.column_dimensions['A'].width=4; ws.column_dimensions['B'].width=8
    ws.column_dimensions['C'].width=34; ws.column_dimensions['D'].width=13
    ws.column_dimensions['E'].width=28; ws.column_dimensions['F'].width=14
    for i in range(7,7+len(months)): ws.column_dimensions[gc(i)].width=10
    cn=7+len(months)
    ws.column_dimensions[gc(cn)].width=12; ws.column_dimensions[gc(cn+1)].width=11
    for i in range(cn+2,cn+7): ws.column_dimensions[gc(i)].width=10
    ws.column_dimensions[gc(cn+7)].width=32
    ws.freeze_panes="D5"

# ════════════════════════════════════════════════════════════════════════════
# SHEET 13: Matrix จำนวนชิ้น – สินค้า × รหัสลูกค้า × เดือน
# ════════════════════════════════════════════════════════════════════════════
def build_sheet13(wb, df, months, prod_names, prod_cat):
    """
    สำหรับแต่ละสินค้า (prod_grp):
      แถว = รหัสลูกค้า
      คอลัมน์ = เดือน (จำนวนชิ้น) + รวม
    แต่ละบล็อกมีหัวสินค้า แล้วตามด้วยรายลูกค้า
    """
    ws = wb.create_sheet("13_Matrix_ชิ้น×ลูกค้า"); ws.sheet_properties.tabColor = "0070C0"

    W(ws,1,1,"📦 Matrix จำนวนชิ้นที่สั่ง — แยกตาม สินค้า × รหัสลูกค้า × เดือน",
      bg="0070C0",fg="FFFFFF",b=True,sz=13,h="left",
      merge_to=3+len(months)+1)
    W(ws,2,1,"💡 แต่ละบล็อก = 1 สินค้า  |  แถว = รหัสลูกค้า  |  คอลัมน์ = เดือน (จำนวนชิ้น)  |  0 = ไม่ได้สั่งเดือนนั้น",
      bg="DDEEFF",fg="003366",sz=8,h="left",merge_to=3+len(months)+1)
    ws.row_dimensions[1].height=26; ws.row_dimensions[2].height=14

    # Sort products by total revenue desc
    prod_rev = df.groupby('prod_grp')['ราคารวม'].sum().sort_values(ascending=False)

    # cond1 for store name display
    code_store = df.groupby('รหัสลูกค้า_norm')['ชื่อร้าน'].agg(lambda x: x.mode()[0])

    ri = 4
    for rank_p, prod_code in enumerate(prod_rev.index, 1):
        pname = prod_names.get(prod_code, prod_code)
        pcat  = prod_cat.get(prod_code, '')

        # Filter data for this product
        pdf = df[df['prod_grp'] == prod_code]
        if len(pdf) == 0: continue

        # Pivot: รหัสลูกค้า_norm × month → จำนวน
        pivot_q = (pdf.groupby(['รหัสลูกค้า_norm','month'])['จำนวน']
                      .sum().unstack(fill_value=0))
        pivot_b = (pdf.groupby(['รหัสลูกค้า_norm','month'])['ราคารวม']
                      .sum().unstack(fill_value=0))
        for m in months:
            if m not in pivot_q.columns: pivot_q[m] = 0
            if m not in pivot_b.columns: pivot_b[m] = 0
        pivot_q = pivot_q[months]; pivot_b = pivot_b[months]
        pivot_q['รวมชิ้น'] = pivot_q.sum(axis=1)
        pivot_b['รวมบาท']  = pivot_b.sum(axis=1)
        # Sort by total qty desc
        pivot_q = pivot_q.sort_values('รวมชิ้น', ascending=False)

        # ── Product header ────────────────────────────────────────────────
        cat_bg = CAT_CLR.get(pcat,'E0E0E0')
        W(ws,ri,1,f"#{rank_p}  {prod_code}",bg="003366",fg="FFFFFF",b=True,sz=10)
        W(ws,ri,2,pname,bg="003366",fg="FFFFFF",b=True,sz=10,h="left",
          merge_to=2+len(months))
        W(ws,ri,3+len(months),pcat,bg=cat_bg,b=True,sz=9)
        W(ws,ri,4+len(months),f"{int(pivot_q['รวมชิ้น'].sum())} ชิ้น",
          bg="0070C0",fg="FFFFFF",b=True,sz=9)
        ws.row_dimensions[ri].height=18; ri+=1

        # ── Column headers ────────────────────────────────────────────────
        W(ws,ri,1,"รหัสลูกค้า",bg="004C97",fg="FFFFFF",b=True,sz=8)
        W(ws,ri,2,"ชื่อร้าน",  bg="004C97",fg="FFFFFF",b=True,sz=8)
        for ci,m in enumerate(months,3):
            W(ws,ri,ci,MTH[m],bg="2E75B6",fg="FFFFFF",b=True,sz=8)
        W(ws,ri,3+len(months),"รวม\n(ชิ้น)",bg="1F4E79",fg="FFFFFF",b=True,sz=8,wrap=True)
        W(ws,ri,4+len(months),"รวม\n(บาท)", bg="1F4E79",fg="FFFFFF",b=True,sz=8,wrap=True)
        ws.row_dimensions[ri].height=22; ri+=1

        # ── Customer rows ─────────────────────────────────────────────────
        for ri_off,(code,qrow) in enumerate(pivot_q.iterrows(),1):
            brow = pivot_b.loc[code] if code in pivot_b.index else pd.Series(dtype=float)
            alt  = "F0F7FF" if ri_off%2==0 else None
            ws.row_dimensions[ri].height=14

            W(ws,ri,1,code,bg=alt,h="left",sz=8)
            W(ws,ri,2,code_store.get(code,''),bg=alt,h="left",sz=8)

            for ci,m in enumerate(months,3):
                qv = int(qrow.get(m,0))
                # Color intensity: more = darker green, 0 = light red
                if qv == 0:
                    cbg = "FFF0F0"; cfg = "CCCCCC"
                elif qv >= pivot_q[m].quantile(0.75):
                    cbg = "C6EFCE"; cfg = "375623"
                elif qv >= pivot_q[m].median():
                    cbg = "E2EFDA"; cfg = "375623"
                else:
                    cbg = alt; cfg = "000000"
                cell = ws.cell(row=ri, column=ci)
                cell.value = qv if qv > 0 else 0
                cell.font = Font(name="Arial", bold=(qv>0), size=9, color=cfg)
                cell.alignment = Alignment(horizontal="center", vertical="center")
                cell.border = BT
                if cbg: cell.fill = F(cbg)
                cell.number_format = '#,##0'

            # Total qty
            tq = int(qrow['รวมชิ้น'])
            W(ws,ri,3+len(months),tq,bg="BDD7EE",b=True,fmt='#,##0')
            # Total baht
            tb = brow.get('รวมบาท', brow[months].sum() if len(brow)>0 else 0)
            W(ws,ri,4+len(months),tb,bg="DEEAF1",fmt='#,##0',sz=8)
            ri+=1

        # ── Product total row ─────────────────────────────────────────────
        W(ws,ri,1,"รวม",bg="1F4E79",fg="FFFFFF",b=True)
        W(ws,ri,2,"",  bg="1F4E79")
        for ci,m in enumerate(months,3):
            W(ws,ri,ci,int(pivot_q[m].sum()),bg="2E75B6",fg="FFFFFF",b=True,fmt='#,##0')
        W(ws,ri,3+len(months),int(pivot_q['รวมชิ้น'].sum()),bg="0070C0",fg="FFFFFF",b=True,fmt='#,##0')
        W(ws,ri,4+len(months),pivot_b['รวมบาท'].sum(),bg="0070C0",fg="FFFFFF",b=True,fmt='#,##0')
        ws.row_dimensions[ri].height=16; ri+=2   # gap between products

    # ── Column widths ─────────────────────────────────────────────────────
    ws.column_dimensions['A'].width = 30
    ws.column_dimensions['B'].width = 26
    for i in range(3, 3+len(months)): ws.column_dimensions[gc(i)].width=10
    ws.column_dimensions[gc(3+len(months))].width=11
    ws.column_dimensions[gc(4+len(months))].width=13
    ws.freeze_panes = "C4"
def build_sheet12(wb, df, grand_total, months, monthly_targets):
    ws=wb.create_sheet("12_เป้าหมายรายเดือน"); ws.sheet_properties.tabColor="FF0000"
    all_months=sorted(set(list(months)+list(monthly_targets.keys())))
    W(ws,1,1,"🎯 เป้าหมายยอดขายรายเดือน vs ยอดจริง",bg="C00000",fg="FFFFFF",b=True,sz=13,h="left",merge_to=8)
    ws.row_dimensions[1].height=26
    total_target=sum(monthly_targets.values())
    actual_in_tgt=sum(df[df['month']==m]['ราคารวม'].sum() for m in monthly_targets)
    ach=actual_in_tgt/total_target if total_target else 0
    for i,(lbl,val,fmt) in enumerate([("เป้าหมายรวม",total_target,'#,##0'),
                                       ("ยอดจริงรวม",actual_in_tgt,'#,##0.00'),
                                       ("% Achievement",ach,'0.0%'),
                                       ("Gap",actual_in_tgt-total_target,'+#,##0;-#,##0;0')]):
        vbg="C6EFCE" if (i==2 and ach>=1) else("FFF2CC" if (i==2 and ach>=0.8) else("FFD7D7" if i==2 else("E2EFDA" if (i==3 and actual_in_tgt>=total_target) else("FFD7D7" if i==3 else "DEEAF1"))))
        W(ws,3,i*2+1,lbl,bg="2E75B6",fg="FFFFFF",b=True,sz=9)
        W(ws,4,i*2+1,val,bg=vbg,b=True,sz=12,fg="1F4E79",fmt=fmt,merge_to=i*2+2)
    ws.row_dimensions[4].height=26
    for ci,h in enumerate(["เดือน","เป้าหมาย (บาท)","ยอดจริง (บาท)","% Achievement","Gap (บาท)","สถานะ","จำนวนร้าน"],1):
        W(ws,6,ci,h,bg="C00000",fg="FFFFFF",b=True,sz=9,wrap=True); ws.row_dimensions[6].height=28
    for ri_off,m in enumerate(all_months,1):
        ri=ri_off+6; actual=df[df['month']==m]['ราคารวม'].sum()
        target=monthly_targets.get(m,None); alt="F9F9F9" if ri_off%2==0 else None
        W(ws,ri,1,MTH.get(m,''),bg=alt,b=True,sz=10)
        W(ws,ri,2,target if target else "—",bg=alt,fmt='#,##0' if target else None,sz=9)
        if m in months:
            W(ws,ri,3,actual,bg=alt,fmt='#,##0.00',sz=9)
            if target:
                a=actual/target; g=actual-target
                ab="C6EFCE" if a>=1 else("FFF2CC" if a>=0.9 else("FCE4D6" if a>=0.7 else "FFD7D7"))
                sl="✅ ถึงเป้า" if a>=1 else("🟡 ใกล้เป้า" if a>=0.9 else("🟠 ต่ำกว่าเป้า" if a>=0.7 else "🔴 ต่ำมาก"))
                W(ws,ri,4,a,bg=ab,b=True,fmt='0.0%',sz=9)
                W(ws,ri,5,g,bg="E2EFDA" if g>=0 else "FFD7D7",fmt='+#,##0;-#,##0;0',sz=9)
                W(ws,ri,6,sl,bg=ab,b=True,sz=9)
                W(ws,ri,7,df[df['month']==m]['ชื่อร้าน'].nunique(),bg=alt,sz=9)
            else:
                for ci in [4,5,6,7]: W(ws,ri,ci,"—",bg="F5F5F5",fg="AAAAAA",sz=8,italic=True)
        else:
            for ci in [3,4,5,6,7]: W(ws,ri,ci,"(ยังไม่มีข้อมูล)",bg="F5F5F5",fg="AAAAAA",sz=8,italic=True)
    for i in range(1,9): ws.column_dimensions[gc(i)].width=15
    ws.freeze_panes="A7"

# ════════════════════════════════════════════════════════════════════════════
# MAIN ENTRY POINT
# ════════════════════════════════════════════════════════════════════════════
def build_dashboard(file_bytes, plc_target=1_250_000, plc_deadline="30 มิถุนายน 2569",
                    recco_target=1_000_000, recco_deadline="31 ธันวาคม 2569",
                    monthly_targets=None, progress_cb=None):
    def upd(pct, msg):
        if progress_cb: progress_cb(pct, msg)

    upd(8,  "📥 โหลดข้อมูล...")
    df, prod_names, prod_cat, targets_from_file = load_data(file_bytes)
    months      = sorted(df['month'].dropna().unique().astype(int))
    grand_total = df['ราคารวม'].sum()
    # เงื่อนไข3 override monthly_targets ถ้ามี
    if targets_from_file:
        monthly_targets = targets_from_file

    wb = Workbook(); wb.remove(wb.active)

    upd(18, "📊 ภาพรวมยอดขาย..."); build_sheet1(wb, df, grand_total, months, monthly_targets)
    upd(28, "🏪 Ranking ร้านค้า..."); build_sheet2(wb, df, grand_total, months)
    upd(36, "🔑 Ranking รหัสลูกค้า..."); build_sheet3(wb, df, grand_total, months)
    upd(43, "🌾 ยอดขายฟาร์ม..."); build_sheet4(wb, df, grand_total, months)
    upd(50, "🎯 เป้าหมาย PLC & Recco..."); build_sheet5(wb, df, grand_total, months, plc_target, plc_deadline, recco_target, recco_deadline)
    upd(57, "📦 ภาพรวมสินค้า (+ ชิ้น)..."); build_sheet6(wb, df, grand_total)
    upd(64, "📋 Ranking สินค้า (+ ชิ้น)..."); build_sheet7(wb, df, grand_total, months, prod_names, prod_cat)
    upd(72, "🔗 สินค้า × Top10 (+ ชิ้น)..."); build_sheets_89(wb, df, grand_total, months, prod_names, prod_cat)
    upd(82, "🔮 Forecast ร้านค้า..."); build_sheet10(wb, df, grand_total, months)
    upd(92, "📦 Forecast สินค้า × ร้าน (+ ชิ้น)..."); build_sheet11(wb, df, grand_total, months, prod_names, prod_cat)
    upd(95, "📊 Matrix ชิ้น × ลูกค้า..."); build_sheet13(wb, df, months, prod_names, prod_cat)
    if monthly_targets:
        upd(96, "🎯 เป้าหมายรายเดือน..."); build_sheet12(wb, df, grand_total, months, monthly_targets)
        upd(98, "📅 ยอดขายรายสัปดาห์..."); build_sheet14(wb, df, months, monthly_targets)

    upd(99, "💾 บันทึกไฟล์...")
    out = io.BytesIO(); wb.save(out); out.seek(0)
    return out.read()

# ════════════════════════════════════════════════════════════════════════════
# SHEET 14: ยอดขายรายสัปดาห์
# ════════════════════════════════════════════════════════════════════════════
def build_sheet14(wb, df, months, monthly_targets):
    ws = wb.create_sheet("14_update รายสัปดาห์")
    ws.sheet_properties.tabColor = "00B050"

    # Week ranges per month (days in each week slot)
    WEEK_RANGES = [(1,7),(8,14),(15,21),(22,None)]  # None = end of month

    import calendar

    # Days in month (assume year 2026)
    YEAR = 2026

    T  = Side(style='thin',   color='D0D0D0')
    M  = Side(style='medium', color='595959')
    BT_  = Border(left=T,right=T,top=T,bottom=T)
    BM_  = Border(left=M,right=M,top=M,bottom=M)

    def F_(h): return PatternFill("solid",fgColor=h)
    def W_(ws_,r,c,val,bg=None,fg="000000",b=False,sz=9,fmt=None,h="center",wrap=False,border=None,merge_to=None):
        cc=ws_.cell(row=r,column=c); cc.value=val
        cc.font=Font(name="Arial",bold=b,size=sz,color=fg)
        cc.alignment=Alignment(horizontal=h,vertical="center",wrap_text=wrap)
        cc.border=border or BT_
        if bg: cc.fill=F_(bg)
        if fmt: cc.number_format=fmt
        if merge_to: ws_.merge_cells(start_row=r,start_column=c,end_row=r,end_column=merge_to)
        return cc

    # Title
    W_(ws,1,1,"📅 [PETSHOP] SALE by WEEKLY",bg="00B050",fg="FFFFFF",b=True,sz=14,h="left",merge_to=6)
    ws.row_dimensions[1].height=26

    ri = 3
    for m in sorted(set(list(months) + list(monthly_targets.keys()))):
        target = monthly_targets.get(m, 0)
        days_in_month = calendar.monthrange(YEAR, m)[1]

        # Month header
        W_(ws,ri,1,f"{MTH[m]} {YEAR}",bg="00B050",fg="FFFFFF",b=True,sz=12,h="left",merge_to=6)
        ws.row_dimensions[ri].height=20; ri+=1

        # KPI row: Target / Actual / %Achieve
        actual_total = df[df['month']==m]['ราคารวม'].sum() if m in months else 0
        ach = actual_total/target if target else 0
        ach_bg = "C6EFCE" if ach>=1 else("FFF2CC" if ach>=0.8 else("FFD7D7" if actual_total>0 else "F5F5F5"))

        W_(ws,ri,1,"Target",bg="E2EFDA",b=True,sz=9)
        W_(ws,ri,2,target,bg="E2EFDA",b=True,fmt='#,##0',sz=10,merge_to=3)
        W_(ws,ri,4,"Actual",bg=ach_bg,b=True,sz=9)
        W_(ws,ri,5,actual_total,bg=ach_bg,b=True,fmt='#,##0.00',sz=10)
        W_(ws,ri,6,ach if target else "—",bg=ach_bg,b=True,fmt='0.0%' if target else None,sz=10)
        ws.row_dimensions[ri].height=18; ri+=1

        # Column headers
        for ci,h_ in enumerate(["Week","Period","ยอดขาย (บาท)","% Contribution\n(vs Target)","% Contribution\n(vs Actual)","สะสม (บาท)"],1):
            W_(ws,ri,ci,h_,bg="1F4E79",fg="FFFFFF",b=True,sz=8,wrap=True)
        ws.row_dimensions[ri].height=32; ri+=1

        # Week rows
        cumulative = 0
        for week_num,(d_start,d_end) in enumerate(WEEK_RANGES,1):
            if d_end is None: d_end = days_in_month
            period_label = f"{d_start}-{d_end}"

            # Sum sales in this week
            if m in months:
                mdf = df[df['month']==m].copy()
                mdf['day'] = mdf['date'].dt.day
                week_sales = mdf[(mdf['day']>=d_start)&(mdf['day']<=d_end)]['ราคารวม'].sum()
            else:
                week_sales = 0

            cumulative += week_sales
            pct_target = week_sales/target if target else 0
            pct_actual = week_sales/actual_total if actual_total else 0
            alt = "F0FFF0" if week_num%2==0 else None

            ws.row_dimensions[ri].height=16
            W_(ws,ri,1,week_num,bg=alt,b=True)
            W_(ws,ri,2,period_label,bg=alt)
            W_(ws,ri,3,week_sales,bg=alt,fmt='#,##0.00',sz=9)
            W_(ws,ri,4,pct_target if target else "—",
               bg="E2EFDA" if pct_target>=0.25 else("FFF2CC" if pct_target>0 else alt),
               fmt='0.0%' if target else None,sz=9)
            W_(ws,ri,5,pct_actual if actual_total else "—",
               bg=alt,fmt='0.0%' if actual_total else None,sz=9)
            W_(ws,ri,6,cumulative,bg="DEEAF1",fmt='#,##0.00',sz=9)
            ri+=1

        # Month total row
        ws.row_dimensions[ri].height=16
        W_(ws,ri,1,"รวม",bg="00B050",fg="FFFFFF",b=True)
        W_(ws,ri,2,"",bg="00B050")
        W_(ws,ri,3,actual_total,bg="00B050",fg="FFFFFF",b=True,fmt='#,##0.00')
        W_(ws,ri,4,ach if target else "—",bg="00B050",fg="FFFFFF",b=True,fmt='0.0%' if target else None)
        W_(ws,ri,5,1.0 if actual_total else "—",bg="00B050",fg="FFFFFF",b=True,fmt='0.0%' if actual_total else None)
        W_(ws,ri,6,actual_total,bg="00B050",fg="FFFFFF",b=True,fmt='#,##0.00')
        ri+=2  # gap

    ws.column_dimensions['A'].width=8
    ws.column_dimensions['B'].width=12
    ws.column_dimensions['C'].width=16
    ws.column_dimensions['D'].width=18
    ws.column_dimensions['E'].width=18
    ws.column_dimensions['F'].width=16
    ws.freeze_panes="A3"
